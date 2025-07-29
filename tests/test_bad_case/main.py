import openpyxl
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
from keye_vl_utils import process_vision_info
from PIL import Image, ImageDraw
from PIL import Image
import torch
import sys
from transformers import AutoTokenizer, AutoModel, AutoProcessor
import contextlib
import os
import shutil
import json

def model_forward(prompt, image_url):
    """模拟模型处理函数，实际使用时替换为真实模型调用"""
    # TODO: 替换为实际的模型调用代码
    return f"模拟响应: {prompt[:20]}... {image_url[-10:]}"


def set_seed(seed: int):
    import random
    import numpy as np

    """设置所有可能的随机数种子，保证实验可重复性"""
    # 设置 Python 内置的随机数种子
    random.seed(seed)
    # 设置 NumPy 的随机数种子
    np.random.seed(seed)
    # 设置 PyTorch 的 CPU 随机数种子
    torch.manual_seed(seed)
    # 设置 PyTorch 的 CUDA 随机数种子（用于 GPU 计算）
    torch.cuda.manual_seed(seed)
    # 如果使用了多个 GPU，还需要设置这个
    torch.cuda.manual_seed_all(seed)
    # 禁用 CuDNN 的非确定性算法（确保结果可复现）
    torch.backends.cudnn.deterministic = True
    # 禁用 CuDNN 的自动调优功能（确保每次运行使用相同的算法）
    torch.backends.cudnn.benchmark = False

set_seed(99999999)


local_rank = 0
model_dir = "/mmu_mllm_hdd_2/zhouyang12/output/Keye/Stage3_0.3.4_1pes_2e-5_resume3k/0.8.0/8b/step3000/global_step3000/converted_hf"
# model_dir = "/mmu_mllm_hdd_2/lingzhixin/release/20250526_hf/"
model = AutoModel.from_pretrained(
model_dir,
torch_dtype=torch.bfloat16,
_attn_implementation = 'flash_attention_2',
device_map="cuda:0",
# ignore_mismatched_sizes=True,
trust_remote_code=True)
processor = AutoProcessor.from_pretrained(model_dir, trust_remote_code=True)
tokenizer = processor.tokenizer
generate_config = {
    "do_sample": False,
    "max_length": 256,
    "top_p": 0.9,
    "num_beams": 1,
    "top_k": 1,
    "temperature": 0.01,
}

generate_config = {
    "do_sample": True,
    "max_length": 256,
    "top_p": 0.9,
    "num_beams": 1,
    "top_k": 1,
    "temperature": 0.01,
}
# generate_config = {}
if 1: generate_config = {
    "do_sample": True,
    "max_length": 256,
    "top_p": 0.95,
    # "top_k": 1,
    "top_k": 20,
    "temperature": 0.6,
}
# generate_config = {}
tag = "081_new_nogreed_sheet1_v2"
sheet_name = "Sheet1"

# /mmu_mllm_hdd_2/zhouyang12/release/20250526

def model_forward(prompt, image_url):
    print(f"forwarding ... ")
    print(f"prompt=\n{prompt}")
    print(f"image_url=\n{image_url}")
    print(f"generate_config=\n{generate_config}")

    mm = [{"type": "image", "image": image_url }]
    messages = [
        {
            "role": "user",
            "content": [
                *mm,
                {"type": "text", "text": prompt},
            ],
        }
    ]
    
    import copy
    text = processor.apply_chat_template(
        messages, tokenize=False, add_generation_prompt=False
    )
    image_inputs, video_inputs = process_vision_info(messages)
    inputs = processor(
        text=[text],
        images=image_inputs,
        videos=video_inputs,
        padding=True,
        return_tensors="pt",
    )
    inputs = inputs.to(local_rank)
    generated = model.generate(**inputs, **generate_config, max_new_tokens=256) 
    output_ids = generated[0][len(inputs.input_ids[0]):].tolist() 
    content = tokenizer.decode(output_ids[0:], skip_special_tokens=True).strip("\n")
    return content


def parse_xlsx(file_path, output_path=None):
    # 生成带时间戳的输出文件名
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.path.dirname(__file__),
            f"processed_results_{tag}_{timestamp}.xlsx"
        )
    # 加载工作簿并获取指定sheet
    wb = openpyxl.load_workbook(file_path)
    try:
        sheet = wb[sheet_name]
    except KeyError:
        print(f"Error: Sheet 'Sheet3' not found in {file_path}")
        print(f"Available sheets: {wb.sheetnames}")
        return
    
    # 创建新工作簿保存结果
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "处理结果"
    
    # 复制表头并添加新列
    headers = [cell.value for cell in sheet[1]] + ["model_response"]
    for col, header in enumerate(headers, 1):
        output_sheet.cell(row=1, column=col, value=header)
    
    # 处理每行数据
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        # 打印原始数据
        for header, value in zip(headers[:-1], row):
            print(f"{header}: {value}")
        
        # 获取prompt和image_url
        missing_fields = []
        if "prompt" not in headers:
            missing_fields.append("prompt")
        else:
            prompt = row[headers.index("prompt")]
            
        if "image_url" not in headers:
            missing_fields.append("image_url")
        else:
            image_url = row[headers.index("image_url")]
        
        if missing_fields:
            print(f"警告: 缺少必要字段 {missing_fields}")
            print("当前行数据:", dict(zip(headers[:-1], row)))
            model_response = "SKIPPED: 缺少必要字段"
        else:
            # 调用模型处理
            model_response = model_forward(prompt, image_url)
            print(f"model_response: {model_response}")
        print("-" * 30)
        
        # 写入新行数据
        for col, value in enumerate(row, 1):
            output_sheet.cell(row=row_idx, column=col, value=value)
        output_sheet.cell(row=row_idx, column=len(headers), value=model_response)
    
    # 保存结果
    output_wb.save(output_path)
    print(f"处理结果已保存到: {output_path}")

if __name__ == "__main__":
    # 替换为实际的xlsx文件路径
    xlsx_file = "/llm_reco/lingzhixin/recovlm_qw0510/recovlm/tests/test_bad_case/reproduce_forward.xlsx"
    parse_xlsx(xlsx_file)
